from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.http.response import JsonResponse
from django.utils.timezone import localtime
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from .youtube_bundle import YoutubeBundle
from .models import *
from .forms import CreateUserForm
import stripe
import datetime
import openpyxl

# Create your views here.
def signup(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        form = CreateUserForm()
        if request.method == "POST":
            form = CreateUserForm(request.POST)
            if form.is_valid():
                form.save()
                user = form.cleaned_data.get('username')
                messages.success(request, 'アカウントが作成されました')
                return redirect('login')
        context = {'form': form}
        return render(request, 'statsgetter/signup.html', context)

def login(request):
    if request.user.is_authenticated:
        return redirect('home')
    else:
        if request.method == "POST":
            username = request.POST['username']
            password = request.POST['password']

            user = authenticate(request, username=username, password=password)

            if user is not None:
                auth_login(request, user)
                return redirect('home')
            else:
                messages.info(request, 'ユーザーネームかパスワードが間違っています')
        context = {}
        return render(request, 'statsgetter/login.html', context)

def logout(request):
    auth_logout(request)
    return redirect('login')

def price(request):
    context = {}
    return render(request, 'statsgetter/price.html', context)

def home(request):
    channels_stats = []
    nicknames = []
    no_such_element_errors = []
    if request.method == "POST":
        # get searched channel titles and make a list of channel titles
        channel_titles_str = request.POST['search']
        channel_titles = channel_titles_str.split('、')

        # when users first visit the website
        # this prevents the users to wait when they first visit the website
        if not channel_titles:
            pass
        # when users search
        else:
            try:
                stats_last = Statistics.objects.latest('id')
                last_transaction_id = stats_last.transaction_id
            except:
                last_transaction_id = 0

            # get stats for the serached channel titles
            for channel_title in channel_titles:
                yt = YoutubeBundle(channel_title)
                results = yt.get_stats_bundle()
                if results == None:
                    no_such_element_errors.append("「" + channel_title + "」" + "でチャンネルが見つかりませんでした。")
                else:
                    for result in results:
                        registered_date = iso_to_python_dt(result["snippet"]["publishedAt"])
                        registered_date = registered_date.strftime("%Y-%m-%d")
                        channel_stats = {
                            'title' : result["snippet"]["title"],
                            'id' : result["id"],
                            'registered_date' : registered_date,
                            'subscriberCount' : int(result["statistics"]["subscriberCount"]),
                            'viewCount' : int(result["statistics"]["viewCount"]),
                            'videoCount' : int(result["statistics"]["videoCount"]),
                        }
                        channels_stats.append(channel_stats)

                        # store channel data in the database
                        youtube_channel = YoutubeChannel.objects.get_or_create(
                            title=channel_stats["title"],
                            channel_id=channel_stats["id"]
                            )

                        # store stats data in the database
                        new_stats = Statistics(
                            subscriber_count=channel_stats["subscriberCount"],
                            view_count=channel_stats["viewCount"],
                            video_count=channel_stats["videoCount"],
                            channel_registered=registered_date
                            )

                        channel = YoutubeChannel.objects.get(channel_id=channel_stats["id"])
                        new_stats.channel = channel

                        new_stats.transaction_id = last_transaction_id + 1
                        new_stats.save()

                        # store nickname data in the database
                        if channel_title != channel_stats["title"]:
                            if Nickname.objects.filter(nickname=channel_title).exists():
                                pass
                            else:
                                nickname = Nickname.objects.create(
                                    nickname = channel_title,
                                    channel = channel
                                )
                                # append to the list to show on the home page
                                nicknames.append(channel_title)

    print(nicknames)
    context = {'channels_stats': channels_stats, 'nicknames': nicknames, 'no_such_element_errors': no_such_element_errors}
    return render(request, 'statsgetter/home.html', context)

@login_required(login_url='login')
def export_excel(request):
    # django setting to export excel
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Youtube_Stats_' + str(datetime.date.today()) + '.xlsx'

    # open excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # only get the latest data
    stats_last = Statistics.objects.latest('id')
    last_transaction_id = stats_last.transaction_id
    stats_latest = Statistics.objects.filter(transaction_id=last_transaction_id)

    # the 1st row is for title
    ws['A1'] = 'チャンネル名'
    ws['B1'] = 'チャンネル登録日'
    ws['C1'] = 'チャンネル登録者数'
    ws['D1'] = '総再生数'
    ws['E1'] = '総動画数'
    ws['F1'] = 'データ取得日'

    # store data
    row = 2
    for data in stats_latest:
        # adjust datetime to excel format
        channel_registered_excel = data.channel_registered.strftime('%m/%d/%Y')
        date_added_excel = localtime(data.date_added).strftime('%m/%d/%Y %H:%M')

        # store data
        ws.cell(row=row, column=1).value = data.channel.title
        ws.cell(row=row, column=2).value = channel_registered_excel
        ws.cell(row=row, column=3).value = data.subscriber_count
        ws.cell(row=row, column=4).value = data.view_count
        ws.cell(row=row, column=5).value = data.video_count
        ws.cell(row=row, column=6).value = date_added_excel
        row += 1

    wb.save(response)

    return response

@csrf_exempt
def stripe_config(request):
    if request.method == 'GET':
        stripe_config = {'publicKey': settings.STRIPE_PUBLISHABLE_KEY}
        return JsonResponse(stripe_config, safe=False)

@csrf_exempt
def create_checkout_session(request):
    if request.method == 'GET':
        domain_url = 'http://localhost:8000/'
        stripe.api_key = settings.STRIPE_SECRET_KEY
        try:
            # Create new Checkout Session for the order
            # Other optional params include:
            # [billing_address_collection] - to display billing address details on the page
            # [customer] - if you have an existing Stripe Customer ID
            # [payment_intent_data] - capture the payment later
            # [customer_email] - prefill the email input in the form
            # For full details see https://stripe.com/docs/api/checkout/sessions/create

            # ?session_id={CHECKOUT_SESSION_ID} means the redirect will have the session ID set as a query param
            checkout_session = stripe.checkout.Session.create(
                success_url = domain_url + 'success?session_id={CHECKOUT_SESSION_ID}',
                cancel_url = domain_url + 'cancelled/',
                payment_method_types = ['card'],
                mode='payment',
                line_items = [
                    {
                        'name': 'Youtube Analyzer Tool',
                        'quantity': 1,
                        'currency': 'jpy',
                        'amount': '1000',
                    }
                ]
            )
            return JsonResponse({'sessionId': checkout_session['id']})
        except Exception as e:
            return JsonResponse({'error': str(e)})

def success(request):
    return render(request, 'statsgetter/success.html')

def cancelled(request):
    return render(request, 'statsgetter/cancelled.html')

def iso_to_python_dt(iso_str):
    dt = None
    try:
        # for the newer videos which have microseconds
        dt = datetime.datetime.strptime(iso_str, "%Y-%m-%dT%H:%M:%S.%fZ")
    except ValueError:
        try:
            # for the older videos which do not have microseconds
            dt = registered_date = datetime.datetime.strptime(iso_str, "%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            pass
    return dt
